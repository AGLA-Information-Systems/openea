import os
from django.test import TestCase
from ..plugins.excel import ExcelPlugin, concept_name_from_excel, relation_name_from_excel
from ..plugins.essential import EssentialPlugin, ESSENTIAL_ONTOLOGY
from ..controllers.knowledge_base import KnowledgeBaseController, BASIC_ONTOLOGY, EXAMPLE_ONTOLOGY, EXAMPLE_INSTANCES
from ..models import OModel, OConcept, OPredicate
from openpyxl import load_workbook


__author__ = "Patrick Agbokou"
__copyright__ = "Copyright 2021, OpenEA"
__credits__ = ["Patrick Agbokou"]
__license__ = "Apache License 2.0"
__version__ = "0.1.0"
__maintainer__ = "Patrick Agbokou"
__email__ = "patrick.agbokou@aglaglobal.com"
__status__ = "Development"


class EssentialPluginTestCase(TestCase):
    def setUp(self):
        pass

    def test_import_ontology(self):
        path='/mnt/c/Data/tmp'
        filename='essential_baseline_v6_11.1.pont'
        model = OModel.get_or_create(name='Metamodel de Test')
        KnowledgeBaseController.ontology_from_dict(model=model, data=ESSENTIAL_ONTOLOGY)
        EssentialPlugin.import_ontology(model=model, path=path, filename=filename)

    def test_import_instances(self):
        self.test_import_ontology()
        path='/mnt/c/Data/tmp'
        filename='essential_baseline_v6_11.1.pins'
        model = OModel.get_or_create(name='Metamodel de Test')
        EssentialPlugin.import_instances(model=model, path=path, filename=filename)

    def test_parse_pont(self):
        data = """
; Tue Apr 20 23:17:11 EDT 2021
; 
;+ (version "3.5")
;+ (build "Build 663")


(defclass %3ACLIPS_TOP_LEVEL_SLOT_CLASS "Fake class to save top-level slot information"
    (is-a USER)
    (role abstract)
    (single-slot sq_index
;+        (comment "An index value for a service quality, allowing them to be sequenced")
        (type INTEGER)
;+        (cardinality 0 1)
        (create-accessor read-write))
    (single-slot rc_revenue_currency
;+        (comment "The currency in which the revenue is generated")
        (type INSTANCE)
;+        (allowed-classes Currency)
;+        (cardinality 0 1)
        (create-accessor read-write))
    (multislot provides_application_services
;+        (comment "Specifies the Application Services that are provided by an Application Provider")
        (type INSTANCE)
;+        (allowed-classes Application_Provider_Role)
;+        (inverse-slot role_for_application_provider)
        (create-accessor read-write)))

(defclass Business_Role_Type "Business Role Types are used to define a taxonomy for Business Roles."
    (is-a Business_Conceptual)
    (role concrete)
    (multislot provides_business_capabilities
;+        (comment "The set of business capabilities that are provided by a role type")
        (type INSTANCE)
;+        (allowed-classes Business_Capability)
        (create-accessor read-write))
    (multislot is_realised_by_roles
;+        (comment "The set of Business Roles that realise a Business Role Type.")
        (type INSTANCE)
;+        (allowed-classes Business_Role)
        (create-accessor read-write))
    (single-slot supersedes_version
;+        (comment "Slot to capture the instance of which this instance is a new version. i.e. Use this slot to define the instance that represents the previous version of the element being considered. Allowed classes must always be a single class of the same type as the domain. e.g. a Technology_Product can only supersede another Technology_Product.")
        (type INSTANCE)
;+        (allowed-classes Business_Role_Type)
;+        (cardinality 0 1)
        (create-accessor read-write)))
"""
        res = EssentialPlugin.parse_pont(data)
        print(res)
        #self.assertEqual(res[1][2][3][1], "fourth long string")

class KnowledgeBaseControllerTestCase(TestCase):
    def setUp(self):
        pass

    def test_sanity(self):
        repository = KnowledgeBaseController.create_repository(name='Metamodel de Test')
        model = KnowledgeBaseController.create_model(repository=repository, name='Metamodel de Test', version='2.3')
        KnowledgeBaseController.ontology_from_dict(model=model, data=ESSENTIAL_ONTOLOGY)

        slot_system_class = KnowledgeBaseController.get_concept(model=model, name=':SLOT')
        kinds = KnowledgeBaseController.get_kinds(model=model, name=':ESSENTIAL-SLOT')
        self.assertIn(slot_system_class, kinds)

        slot_templates = KnowledgeBaseController.get_slot_templates(model=model, name=':SLOT')
        slot_template_names =  [item.name for sublist in [x[1] for x in slot_templates] for item in sublist]
        print(slot_template_names)
        self.assertIn('name', slot_template_names)

        class_name = 'Test Suite'
        class_concept = KnowledgeBaseController.create_concept(model=model, name=class_name, description='')
        found_class_concept = KnowledgeBaseController.get_concept(model=model, name=class_name)
        self.assertEqual(class_concept, found_class_concept)

        instance_name = 'Test Case'
        instance_concept = KnowledgeBaseController.create_instance(model=model, name=instance_name, concept=class_concept)
        found_instance_concept = KnowledgeBaseController.get_concept(model=model, name=instance_name)
        self.assertEqual(instance_concept, found_instance_concept)

        class_name = 'Test Suite 2'
        class_concept = KnowledgeBaseController.create_concept(model=model, name=class_name, description='')

        instance_concept_2 = KnowledgeBaseController.create_instance(model=model, name=instance_name, concept=class_concept)
        found_instance_concept_2 = KnowledgeBaseController.get_concept(model=model, name=instance_name, class_name=class_name)
        self.assertNotEqual(found_instance_concept, found_instance_concept_2)

        class_name = 'Test Suite'
        found_instance_concept_3 = KnowledgeBaseController.get_concept(model=model, name=instance_name, class_name=class_name)
        self.assertEqual(found_instance_concept, found_instance_concept_3)

        #slot_name = 'Class Test'
        #slot_name = KnowledgeBaseController.create_slot(model=model, slot_name=slot_name, instance_name=instance_name, concept=class_concept)

    def test_import_essential_export_excel_instances(self):
        model = OModel.get_or_create(name='Metamodel de Test')
        KnowledgeBaseController.ontology_from_dict(model=model, data=ESSENTIAL_ONTOLOGY)
        path='/mnt/c/Data/tmp'
        filename='essential_baseline_v6_11.1.pont'
        EssentialPlugin.import_ontology(model=model, path=path, filename=filename)
        path='/mnt/c/Data/tmp'
        filename='essential_baseline_v6_11.1.pins'
        model = OModel.get_or_create(name='Metamodel de Test')
        EssentialPlugin.import_instances(model=model, path=path, filename=filename)
        path='/mnt/c/Data/tmp'
        filename='essential_instances.xlsx'
        ExcelPlugin.export_instances(model=model, path=path, filename=filename)

class ExcelPluginTestCase(TestCase):
    def setUp(self):
        pass

    def test_export_ontology(self):
        path='/mnt/c/Data/tmp'
        filename='ontology.xlsx'

        repository = KnowledgeBaseController.create_repository(name='Metamodel de Test')
        model = KnowledgeBaseController.create_model(repository=repository, name='Metamodel de Test', version='2.3')
        KnowledgeBaseController.ontology_from_dict(model=model, data=ESSENTIAL_ONTOLOGY)

        ExcelPlugin.export_ontology(model=model, path=path, filename=filename)
        result = False
        wb = load_workbook(os.path.join(path, filename), read_only=True)
        ws = wb['entities']
        for row in ws.rows:
            for cell in row:
                if cell.value == ':FACET':
                    result = True
        wb.close()
        self.assertTrue(result)

    def test_import_ontology(self):
        path='/mnt/c/Data/tmp'
        filename='ontology.xlsx'
        
        repository = KnowledgeBaseController.create_repository(name='Metamodel de Test')
        model = KnowledgeBaseController.create_model(repository=repository, name='Metamodel de Test', version='2.3')

        ExcelPlugin.import_ontology(model=model, path=path, filename=filename)
        slot_master = KnowledgeBaseController.get_concept(model=model, name=':SLOT')
        self.assertEqual(slot_master.name, ':SLOT')

    def test_export_instances(self):
        path='/mnt/c/Data/tmp'
        filename = 'instances.xlsx'

        repository = KnowledgeBaseController.create_repository(name='Metamodel de Test')
        model = KnowledgeBaseController.create_model(repository=repository, name='Metamodel de Test', version='2.3')
        KnowledgeBaseController.ontology_from_dict(model=model, data=BASIC_ONTOLOGY)
        KnowledgeBaseController.ontology_from_dict(model=model, data=EXAMPLE_ONTOLOGY)
        KnowledgeBaseController.instances_from_dict(model=model, data=EXAMPLE_INSTANCES)

        ExcelPlugin.export_instances(model=model, path=path, filename=filename)
        result = False
        wb = load_workbook(os.path.join(path, filename), read_only=True)
        ws = wb['instances']
        for row in ws.rows:
            for cell in row:
                if cell.value == 'Kossi':
                    result = True
        wb.close()
        self.assertTrue(result)

    def test_import_instances(self):
        path='/mnt/c/Data/tmp'
        filename = 'instances.xlsx'
        
        repository = KnowledgeBaseController.create_repository(name='Metamodel de Test')
        model = KnowledgeBaseController.create_model(repository=repository, name='Metamodel de Test', version='2.3')
        KnowledgeBaseController.ontology_from_dict(model=model, data=BASIC_ONTOLOGY)
        KnowledgeBaseController.ontology_from_dict(model=model, data=EXAMPLE_ONTOLOGY)

        ExcelPlugin.import_instances(model=model, path=path, filename=filename)
        kossi = KnowledgeBaseController.get_concept(model=model, name='Kossi')
        slot_templates = KnowledgeBaseController.get_slot_templates(model=model, name='')
        print(slot_templates)
        slots = KnowledgeBaseController.get_slots(model=model, instance=kossi)
        print(slot_templates)
