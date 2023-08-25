import os
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from .utils import name_from_excel, name_to_excel, parse_pont
from .models import OClass, OSlot, OFacet, OInstance
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import urllib.parse


class KnowledgeBaseController:

    def import_excel(path, filename='import.xlsx'):
        
        with transaction.atomic():
            wb = load_workbook(os.path.join(path, filename), read_only=True)
            sheet_names = wb.get_sheet_names()

            if 'classes' in sheet_names:
                ws = wb['classes']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    name = name_from_excel(ws['A' + str(index)].value)
                    description = ws['B' + str(index)].value
                    
                    oclass_parent_names = ws['C' + str(index)].value or []
                    if oclass_parent_names:
                        oclass_parent_names = [name_from_excel(x) for x in oclass_parent_names.split(',')]

                    print(index, ' ',name, ' ', description, ' ', oclass_parent_names)
                    oclass_parents = []
                    try:
                        oclass_parents = [OClass.objects.get(name=x) for x in oclass_parent_names]
                    except ObjectDoesNotExist:
                        pass

                    try:
                        oclass = OClass.objects.get(name=name)
                        oclass.description = description
                        oclass.is_a.set(oclass_parents)
                        oclass.save()
                    except ObjectDoesNotExist:
                        oclass = OClass.objects.create(name=name, description=description)
                        oclass.is_a.set(oclass_parents)
                        oclass.save()
                    
                    
            if 'slots' in sheet_names:
                ws = wb['slots']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    name = name_from_excel(ws['A' + str(index)].value)
                    description = ws['B' + str(index)].value
                    oclass_name = name_from_excel(ws['C' + str(index)].value)

                    print(index, ' ',name, ' ', description, ' ', oclass_name)
                    try:
                        oclass = OClass.objects.get(name=oclass_name)
                        try:
                            oslot = OSlot.objects.get(name=name)
                            oslot.description = description
                            oslot.save()
                        except ObjectDoesNotExist:
                            oslot = OSlot.objects.create(name=name, description=description, oclass=oclass)
                            oslot.save()
                    except ObjectDoesNotExist:
                        raise ValueError('MISSING_OCLASS:' + oclass_name)
            
            if 'facets' in sheet_names:
                ws = wb['facets']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    name = name_from_excel(ws['A' + str(index)].value)
                    description = ws['B' + str(index)].value
                    oslot_name = name_from_excel(ws['C' + str(index)].value)

                    print(index, ' ', name, ' ', description, ' ', oslot_name)
                    try:
                        oslot = OSlot.objects.get(name=oslot_name)
                        try:
                            ofacet = OFacet.objects.get(name=name)
                            ofacet.description = description
                            ofacet.save()
                        except ObjectDoesNotExist:
                            ofacet = OFacet.objects.create(name=name, description=description, oslot=oslot)
                            ofacet.save()
                    except ObjectDoesNotExist:
                        raise ValueError('MISSING_OSLOT:' + oslot_name)

            if 'instances' in sheet_names:
                ws = wb['instances']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    name = name_from_excel(ws['A' + str(index)].value)
                    description = ws['B' + str(index)].value
                    oclass_name = name_from_excel(ws['C' + str(index)].value)

                    print(index, ' ', name, ' ', description, ' ', oclass_name)
                    try:
                        oclass = OInstance.objects.get(name=oclass_name)
                        try:
                            oinstance = OInstance.objects.get(name=name)
                            oinstance.description = description
                            oinstance.save()
                        except ObjectDoesNotExist:
                            oinstance = OInstance.objects.create(name=name, description=description, oclass=oclass)
                            oinstance.save()
                    except ObjectDoesNotExist:
                        raise ValueError('MISSING_OCLASS:' + oslot_name)

            # Close the workbook after reading
            wb.close()


    def export_excel(path, filename='export.xlsx'):
        wb = Workbook()
        
        sheet_names = ['classes', 'slots', 'facets']
        for sheet_name in sheet_names:
            wb.create_sheet(sheet_name)
        # for sheet_name in wb.get_sheet_names():
        #     if sheet_name not in sheet_names:
        #         wb.remove_sheet(sheet_name)

        # Write classes
        sheet = wb['classes']
        sheet.append(['Name', 'Description', 'Kind'])
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 25
        #StyleController.apply_header_style(sheet["A1"])
        #StyleController.apply_header_style(sheet["B1"])
        for oclass in OClass.objects.all().order_by('name'):
            parent_class_names = [x.name for x in oclass.is_a.all()]
            sheet.append([name_to_excel(oclass.name), oclass.description, ','.join([name_to_excel(x) for x in parent_class_names])])
        table = Table(displayName="classes", ref="A1:C{}".format(OClass.objects.count() + 2))
        StyleController.apply_table_style(table)
        sheet.add_table(table)

        # Write slots
        sheet = wb['slots']
        sheet.append(['Name', 'Description', 'Class'])
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 25
        for oslot in OSlot.objects.all().order_by('name'):
            sheet.append([name_to_excel(oslot.name), oslot.description, name_to_excel(oslot.oclass.name)])
        table = Table(displayName="slots", ref="A1:C{}".format(OSlot.objects.count() + 2))
        StyleController.apply_table_style(table)
        sheet.add_table(table)

        # Write facets
        sheet = wb['facets']
        sheet.append(['Name', 'Description', 'Slot'])
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 25
        for ofacet in OFacet.objects.all().order_by('name'):
            sheet.append([name_to_excel(ofacet.name), ofacet.description, name_to_excel(ofacet.oslot.name)])
        table = Table(displayName="facets", ref="A1:C{}".format(OFacet.objects.count() + 2))
        StyleController.apply_table_style(table)
        sheet.add_table(table)

        sheet = wb['Sheet']
        wb.remove(sheet)
        wb.save(os.path.join(path, filename))

    def import_ontology(path, filename='import.pont'):
        with open(os.path.join(path, filename), 'r') as fimport:
            buffer = []
            count = 0
            line_count = 0
            for line in fimport:
                line_count = line_count + 1
                buffer.append(urllib.parse.unquote(line))
                if line == "\n":
                    count = count + 1
                    data = parse_pont(' '.join(buffer))
                    buffer = []

                    for datum in data:
                        if 'type' in datum:
                            # if len(datum['values']) == 1:
                            #     print(datum['type'], ":", datum['values'][0])
                            if len(datum['values']) > 1:
                                print(datum['type'], ":", datum['values'][0], "=>", datum['values'][1])
                                for subdatum in datum['values']:
                                    print(subdatum)
   
            with transaction.atomic():
                pass
    
    def create_concept(data):
        if data['type'] == 'defclass':
            oclass = OClass.objects.create(name=data['values'][0], description=datum['values'][1])
            for x in data['values'][2:]:
                if x['type'] == 'is-a-subtype-of':
                    oclass.is_a.set([OClass.objects.get(name=x) for x in x['values']])
                if x['type'] == 'role':
                    oclass.role = x['values'][0]
                if x['type'] == 'single-slot':
                    oslot = OClass.get_or_create(name=data['values'][0])

                    for y in data['values'][1:]:
                        if y['type'] == 'type':
                            value_type = y['values'][0]
                        if y['type'] == 'cardinality':
                            value_type = y['values'][0]

                if x['type'] == 'multiple-slot':
                    oclass.role = x['values']


    def export_pont(path, filename='import.pont'):
        raise NotImplementedError

    def import_yaml():
        raise NotImplementedError

    def export_yaml():
        raise NotImplementedError
