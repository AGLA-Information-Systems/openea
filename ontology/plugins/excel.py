import os

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side)
from openpyxl.worksheet.table import Table, TableStyleInfo

from ontology.controllers.knowledge_base import KnowledgeBaseController
from ontology.controllers.o_model import ModelUtils
from ontology.models import OConcept, OInstance, OPredicate, ORelation, OSlot
from ontology.plugins.plugin import ACTION_EXPORT, ACTION_IMPORT, Plugin

__author__ = "Patrick Agbokou"
__copyright__ = "Copyright 2021, OpenEA"
__credits__ = ["Patrick Agbokou"]
__license__ = "Apache License 2.0"
__version__ = "1.0.0"
__maintainer__ = "Patrick Agbokou"
__email__ = "patrick.agbokou@aglaglobal.com"
__status__ = "Development"


class ExcelPlugin(Plugin):
    def available_actions():
        return {ACTION_IMPORT, ACTION_EXPORT}

    def get_format():
        return ('EXCEL', 'Excel')

    def get_file_extension(knowledge_set):
        return 'xlsx'

    def import_ontology(model, path, filename='ontology.xlsx', filters=None):

        with transaction.atomic():
            wb = load_workbook(os.path.join(path, filename), read_only=True)
            sheet_names = wb.get_sheet_names()                        

            if 'concepts' in sheet_names and get_boolean_filter(filters, 'concepts'):
                ws = wb['concepts']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    id = ws['A' + str(index)].value or None
                    name = name_from_excel(ws['B' + str(index)].value)
                    description = ws['C' + str(index)].value or ''
                    if name:
                        oconcept = OConcept.get_or_create(name=name, model=model, description=description, id=None)
                        oconcept.description = description
                        oconcept.save()

                    
            if 'relations' in sheet_names and get_boolean_filter(filters, 'relations'):
                ws = wb['relations']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    id = ws['A' + str(index)].value or None
                    name = name_from_excel(ws['B' + str(index)].value)
                    description = ws['C' + str(index)].value or ''
                    type = ws['D' + str(index)].value
                    if name:
                        relation = ORelation.get_or_create(model=model, name=name, type=type, description=description, id=id)
                        relation.description = description
                        relation.type = type
                        relation.save()


            if 'ontology' in sheet_names and get_boolean_filter(filters, 'predicates'):
                ws = wb['ontology']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    id = ws['A' + str(index)].value or None
                    subject_id = ws['B' + str(index)].value or None
                    subject_name = name_from_excel(ws['C' + str(index)].value)
                    predicate_id = ws['D' + str(index)].value or None
                    predicate_name = name_from_excel(ws['E' + str(index)].value)
                    object_id = ws['F' + str(index)].value or None
                    object_name = name_from_excel(ws['G' + str(index)].value)
                    cardinality_min = ws['H' + str(index)].value or 0
                    cardinality_max = ws['I' + str(index)].value or 0

                    #print(id,' ',subject_id,' ',subject_name,' ',predicate_id,' ',predicate_name,' ',object_id,' ',object_name)
                    if predicate_name:
                        relation = ORelation.get_or_create(model=model, name=predicate_name)
                        subject = OConcept.get_or_create(model=model, name=subject_name, id=subject_id)
                        object = OConcept.get_or_create(model=model, name=object_name, id=object_id)

                        OPredicate.get_or_create(model=model, relation=relation, subject=subject, object=object, cardinality_min=cardinality_min, cardinality_max=cardinality_max)
                 

            # Close the workbook after reading
            wb.close()

    def export_ontology(model, path, filename='ontology.xlsx', filters=None):
        relation_ids = ModelUtils.get_filter(filters, 'relation_ids')
        concept_ids = ModelUtils.get_filter(filters, 'concept_ids')
        predicate_ids = ModelUtils.get_filter(filters, 'predicate_ids')

        wb = Workbook()
        
        sheet_names = ['ontology', 'concepts', 'relations']
        for sheet_name in sheet_names:
            wb.create_sheet(sheet_name)

        # Write ontology
        sheet = wb['ontology']
        sheet.append(['Entry ID', 'Concept ID', 'Concept', 'Relation ID', 'Relation', 'Related Concept ID', 'Related Concept', 'Cardinality Minimal', 'Cardinality Maximal'])
        count = 1
        sheet.column_dimensions['A'].hidden = True

        sheet.column_dimensions['B'].hidden = True
        sheet.column_dimensions['C'].width = 25

        sheet.column_dimensions['D'].hidden = True
        sheet.column_dimensions['E'].width = 25

        sheet.column_dimensions['F'].hidden = True
        sheet.column_dimensions['G'].width = 25

        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10

        predicate_query = OPredicate.objects.filter(model=model)
        if predicate_ids:
            predicate_query = predicate_query.filter(id__in=predicate_ids)
        for predicate in predicate_query.order_by('subject__name').order_by('relation__name').order_by('object__name'):
            count = count + 1
            sheet.append([str(predicate.id),
                        str(predicate.subject.id),
                        name_to_excel(predicate.subject.name),
                        str(relation.id),
                        name_to_excel(relation.name),
                        str(predicate.object.id),
                        name_to_excel(predicate.object.name),
                        predicate.cardinality_min,
                        predicate.cardinality_max])

        table = Table(displayName="ontology", ref="A1:G{}".format(count))
        StyleController.apply_table_style(table)
        sheet.add_table(table)

        # Write concepts
        sheet = wb['concepts']
        sheet.append(['ID', 'Name', 'Description'])
        sheet.column_dimensions['A'].hidden = True
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 60
        count = 1

        concept_query = OConcept.objects.filter(model=model)
        if concept_ids:
            concept_query = concept_query.filter(id__in=concept_ids)
        for concept in concept_query.order_by('name'):
            sheet.append([str(concept.id),
                          name_to_excel(concept.name),
                          concept.description])
            count = count + 1
        sheet.append(['', '', ''])
        count = count + 1

        table = Table(displayName="concepts", ref="A1:C{}".format(count))
        StyleController.apply_table_style(table)
        sheet.add_table(table)

        # Write relations
        sheet = wb['relations']
        sheet.append(['ID', 'Name', 'Description', 'Type'])
        sheet.column_dimensions['A'].hidden = True
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 60
        sheet.column_dimensions['D'].width = 25
        count = 1

        relation_query = ORelation.objects.filter(model=model)
        if relation_ids:
            relation_query = relation_query.filter(id__in=relation_ids)
        for relation in relation_query.order_by('name'):
            sheet.append([str(relation.id),
                          name_to_excel(relation.name),
                          relation.description,
                          relation.type])
            count = count + 1
        sheet.append(['', '', ''])
        count = count + 1

        table = Table(displayName="predicates", ref="A1:C{}".format(count))
        StyleController.apply_table_style(table)
        sheet.add_table(table)

        sheet = wb['Sheet']
        wb.remove(sheet)
        wb.save(os.path.join(path, filename))


    def import_instances(model, path, filename='instances.xlsx', filters=None):
         
        with transaction.atomic():
            wb = load_workbook(os.path.join(path, filename), read_only=True)
            sheet_names = wb.get_sheet_names()                        

            if 'instances' in sheet_names  and get_boolean_filter(filters, 'instances'):
                ws = wb['instances']
                max_row=ws.max_row
                for index in range(2, max_row+1):
                    instance_id = ws['A' + str(index)].value or None
                    instance_name = name_from_excel(ws['B' + str(index)].value)
                    instance_code = name_from_excel(ws['C' + str(index)].value)
                    instance_description = ws['D' + str(index)].value or ''

                    concept_id = ws['E' + str(index)].value or None
                    concept_name = name_from_excel(ws['F' + str(index)].value)
                    
                    slot_id = ws['G' + str(index)].value or None
                    slot_name = name_from_excel(ws['H' + str(index)].value)
                    slot_order  = ws['I' + str(index)].value or '0'
                    slot_description = ws['J' + str(index)].value or ''

                    predicate_id = ws['K' + str(index)].value or None
                    predicate_cardinality_min = ws['L' + str(index)].value or 0
                    predicate_cardinality_max = ws['M' + str(index)].value or 0

                    relation_id = ws['N' + str(index)].value or None
                    relation_id_name = name_from_excel(ws['O' + str(index)].value)

                    object_concept_id = ws['P' + str(index)].value or None
                    object_concept_name = name_from_excel(ws['Q' + str(index)].value)

                    object_id = ws['R' + str(index)].value or None
                    object_name = name_from_excel(ws['S' + str(index)].value)
                    object_code = name_from_excel(ws['T' + str(index)].value)
                    object_description = ws['U' + str(index)].value or ''

                    if instance_name:
                        concept = OConcept.get_or_create(model=model, name=concept_name, id=concept_id)
                        relation = ORelation.get_or_create(model=model, name=relation_id_name, id=relation_id)
                        object_concept = OConcept.get_or_create(model=model, name=object_concept_name, id=object_concept_id)
                        predicate = OPredicate.get_or_create(model=model, relation=relation, subject=concept, object=object_concept, cardinality_min=predicate_cardinality_min, cardinality_max=predicate_cardinality_max, id=predicate_id)
                        
                        instance = OInstance.get_or_create(model=model, name=instance_name, code=instance_code, concept=concept, id=instance_id, description=instance_description)
                        object = OInstance.get_or_create(model=model, name=object_name, code=object_code, concept=object_concept, id=object_id, description=object_description)
                        
                        slot = OSlot.get_or_create(model=model, predicate=predicate, description=slot_description, order=slot_order, subject=instance, object=object, id=slot_id)

            # Close the workbook after reading
            wb.close()

    def export_instances(model, path, filename='instances.xlsx', filters=None):
        predicate_ids = ModelUtils.get_filter(filters, 'predicate_ids')
        instance_ids = ModelUtils.get_filter(filters, 'instance_ids')
        
        wb = Workbook()
        
        sheet_names = ['instances']
        for sheet_name in sheet_names:
            wb.create_sheet(sheet_name)

        # Write instances
        sheet = wb['instances']
        sheet.append(['InstanceID', 'Instance', 'Instance Code', 'Description',
                      'ConceptID', 'Concept',
                      'SlotID', 'Slot' , 'Slot Order', 'Slot Description',
                      'PredicateID', 'Cardinality Min', 'Cardinality Max',
                      'RelationID', 'Relation',
                      'Object ConceptID', 'Object Concept',
                      'ObjectID', 'Object', 'Object Code', 'Object Description'])
        sheet.column_dimensions['A'].hidden = True
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].hidden = True

        sheet.column_dimensions['E'].hidden = True
        sheet.column_dimensions['F'].width = 25

        sheet.column_dimensions['G'].hidden = True
        sheet.column_dimensions['H'].width = 25
        sheet.column_dimensions['I'].width = 25
        sheet.column_dimensions['J'].hidden = True

        sheet.column_dimensions['K'].hidden = True
        sheet.column_dimensions['L'].width = 10
        sheet.column_dimensions['M'].width = 10

        sheet.column_dimensions['N'].hidden = True
        sheet.column_dimensions['O'].width = 25

        sheet.column_dimensions['P'].hidden = True
        sheet.column_dimensions['Q'].width = 25

        sheet.column_dimensions['R'].hidden = True
        sheet.column_dimensions['S'].width = 25
        sheet.column_dimensions['T'].width = 10
        sheet.column_dimensions['U'].hidden = True
        
        count = 1
        instance_query = OInstance.objects.filter(model=model)
        if instance_ids:
            instance_query = instance_query.filter(id__in=instance_ids)

        for instance in instance_query.order_by('name'):
            for slot in OSlot.objects.filter(model=model, subject=instance).all():
                sheet.append([str(instance.id), name_to_excel(instance.name), name_to_excel(instance.code), instance.description,
                              str(instance.concept.id), name_to_excel(instance.concept.name),
                              str(slot.id), name_to_excel(slot.name), slot.order, slot.description,
                              str(slot.predicate.id), str(slot.predicate.cardinality_min), str(slot.predicate.cardinality_max),
                              str(slot.predicate.relation.id),  name_to_excel(slot.predicate.relation.name),
                              str(slot.object.concept.id), name_to_excel(slot.object.concept.name),
                              str(slot.object.id), name_to_excel(slot.object.name), name_to_excel(slot.object.code), slot.object.description])
                count = count + 1

        table = Table(displayName="instances", ref="A1:U{}".format(count))
        StyleController.apply_table_style(table)
        sheet.add_table(table)

        sheet = wb['Sheet']
        wb.remove(sheet)
        wb.save(os.path.join(path, filename))
        
class StyleController:
    @staticmethod
    def apply_default_style(cell):
        cell.font = Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
        cell.fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FFFFFFFF')
        # cell.border = Border(
        #         left=Side(border_style=None, color='FF000000'),
        #         right=Side(border_style=None, color='FF000000'),
        #         top=Side(border_style=None, color='FF000000'),
        #         bottom=Side(border_style=None, color='FF000000'),
        #         diagonal=Side(border_style=None, color='FF000000'),
        #         diagonal_direction=0,
        #         outline=Side(border_style=None, color='FF000000'),
        #         vertical=Side(border_style=None, color='FF000000'),
        #         horizontal=Side(border_style=None, color='FF000000'))
        # cell.alignment = Alignment(
        #         horizontal='general',
        #         vertical='bottom',
        #         text_rotation=0,
        #         wrap_text=False,
        #         shrink_to_fit=False,
        #         indent=0)
        # cell.number_format = 'General',
        # cell.protection = Protection(locked=False, hidden=False)

    @staticmethod
    def apply_header_style(cell):
        style = StyleController.apply_default_style(cell)
        cell.font = Font(
            name='Calibri',
            size=14,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FFFFFFFF')
        cell.fill = PatternFill(fill_type='solid', start_color='AA00b1ac', end_color='AA00b1ac')

    @staticmethod
    def apply_table_style(table):
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

#=================================================================================================
# Useful functions

def name_from_excel(name):
    return name or ''
    if name:
        if name.isupper():
            return name
        return '_'.join([x.capitalize() for x in name.split(' ')])
    return None

def name_to_excel(name):
    return name or ''
    if name:
        return name.replace('_', ' ')
    return None

def get_boolean_filter(filters, varname):
    if varname in filters and not filters[varname]:
        return False
    return True