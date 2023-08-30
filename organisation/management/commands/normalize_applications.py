import json
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from ontology.models import (OConcept, OInstance, OModel, OPredicate,
                             ORelation, OSlot)


class Command(BaseCommand):
    help = 'Create an model with its permissions'

    def add_arguments(self, parser):
        parser.add_argument('model_name', nargs=1, type=str)
        parser.add_argument('model_version', nargs=1, type=str)

    def handle(self, *args, **options):
        model_name = options['model_name'][0]
        model_version = options['model_version'][0]
        
        with transaction.atomic():
            try:
                model, created = OModel.objects.get_or_create(name=model_name, version=model_version)
                print(model.id)
                property_relation, created = ORelation.objects.get_or_create(model=model, name='a pour propriété', defaults={'description': ''})
                utilise_relation, created = ORelation.objects.get_or_create(model=model, name='utilise', defaults={'description': ''})
                implemente_par_relation, created = ORelation.objects.get_or_create(model=model, name='est implémenté par', defaults={'description': ''})
                fait_partie_relation, created = ORelation.objects.get_or_create(model=model, name='fait partie de', defaults={'description': ''})

                etablissement_concept, created = OConcept.objects.get_or_create(model=model, name='Établissement', defaults={'description': ''})
                installation_concept, created = OConcept.objects.get_or_create(model=model, name='Installation', defaults={'description': ''})
                system_concept, created = OConcept.objects.get_or_create(model=model, name='Système informatique ou Outil', defaults={'description': ''})
                system_instance_concept, created = OConcept.objects.get_or_create(model=model, name='Instance de Système informatique', defaults={'description': ''})
                etiquette_concept, created = OConcept.objects.get_or_create(model=model, name='Étiquette', defaults={'description': ''})
                contrat_concept, created = OConcept.objects.get_or_create(model=model, name='Contrat', defaults={'description': ''})
                manufacturier_concept, created = OConcept.objects.get_or_create(model=model, name='Manufacturier', defaults={'description': ''})
                
                installation_etablissement_predicate, created = OPredicate.objects.get_or_create(model=model, subject=installation_concept, relation=fait_partie_relation, object=etablissement_concept, defaults={'description': ''})
                system_etablissement_predicate, created = OPredicate.objects.get_or_create(model=model, subject=system_concept, relation=property_relation, object=etablissement_concept, defaults={'description': ''})
                system_instance_etablissement_predicate, created = OPredicate.objects.get_or_create(model=model, subject=system_instance_concept, relation=property_relation, object=etablissement_concept, defaults={'description': ''})
                system_manufacturier_predicate, created = OPredicate.objects.get_or_create(model=model, subject=system_concept, relation=property_relation, object=manufacturier_concept, defaults={'description': ''})
        
                system_instance_system_predicate, created = OPredicate.objects.get_or_create(model=model, subject=system_instance_concept, relation=property_relation, object=system_concept, defaults={'description': ''})
                system_instance_installation_predicate, created = OPredicate.objects.get_or_create(model=model, subject=system_instance_concept, relation=property_relation, object=installation_concept, defaults={'description': ''})
                system_instance_etiquette_predicate, created = OPredicate.objects.get_or_create(model=model, subject=system_instance_concept, relation=property_relation, object=etiquette_concept, defaults={'description': ''})
                system_etiquette_predicate, created = OPredicate.objects.get_or_create(model=model, subject=system_concept, relation=property_relation, object=etiquette_concept, defaults={'description': ''})
                
                
                processus_concept = OConcept.objects.get(model=model, name='Processus')
                activite_concept = OConcept.objects.get(model=model, name='Activité')
                processus_system_predicate  = OPredicate.objects.get(model=model, subject=processus_concept, relation=property_relation, object=system_concept)
                activite_system_predicate = OPredicate.objects.get(model=model, subject=activite_concept, relation=property_relation, object=system_concept)

                document_dir = Path(r'/mnt/c/Data/cartographie')
                paths = [
                    'applications_normees.xlsx'
                ]
                for path in paths:
                    for xlsx_file in document_dir.glob(path):
                        # xlsx_file is a Path object
                        # if you use old libraries, you have to use str(xlsx_file) to convert the Path to a str
                        print(xlsx_file)

                        workbook = load_workbook(filename=xlsx_file, read_only=True)
                        #sheet = workbook.active
                        for sheet in workbook.worksheets:
                            print(sheet.title)
                            limit = sheet.max_row + 1
                            i = 2
                            data = []
                            while i < limit:
                                nom_system = sheet['C' + str(i)].value
                                nom_flux = sheet['E' + str(i)].value
                                if nom_system and nom_flux and nom_flux == 'SI':
                                    print(nom_system, ': ', i)
                                    system_data = {
                                        'nom': sheet['C' + str(i)].value.strip(),
                                        'etablissement': sheet['B' + str(i)].value.strip() if sheet['B' + str(i)].value else None,
                                        'code': sheet['A' + str(i)].value.strip() if sheet['A' + str(i)].value else None,
                                        'manufacturier': sheet['E' + str(i)].value.strip() if sheet['E' + str(i)].value else None,
                                        'description': sheet['F' + str(i)].value.strip() if sheet['F' + str(i)].value else None
                                    }
                                    data.append(system_data)
                                i = i + 1

                            #print(json.dumps(data, indent=4, ensure_ascii=False))

                            contrat_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='Contrat inconnu', defaults={'description': ''})

                            dsn_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='DSN', defaults={'description': ''})
                            cible_dsn_vitrine_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='Cible DSN Vitrine', defaults={'description': ''})
                            cible_dsn_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='Cible DSN', defaults={'description': ''})
                            analyse_dsn_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='Analyse DSN complétée', defaults={'description': ''})


                            for old_system in OInstance.objects.filter(concept=system_concept):
                                existing_slots = list(OSlot.objects.filter(object=old_system, predicate=processus_system_predicate).all()) + \
                                                 list(OSlot.objects.filter(object=old_system, predicate=activite_system_predicate).all()) + \
                                                 list(OSlot.objects.filter(predicate=system_instance_system_predicate, object=old_system).all())
                                if len(existing_slots) > 0:
                                    print("Système à garder : {}".format(old_system.name))
                                else:
                                    old_system.delete()


                            for system_data in data:
                                #####
                                similar_systems = OInstance.objects.filter(name__icontains=system_data['nom'], concept=system_concept)
                                similar_system_instances = OInstance.objects.filter(name__icontains=system_data['nom'], concept=system_instance_concept)

                                print("Système normalisé : {}".format(system_data['nom']))
                                print("Système similaires : {}".format([x.name for x in similar_systems]))
                                print("Intances de système trouvés : {}".format([x.name for x in similar_system_instances]))

                                system, created = OInstance.objects.update_or_create(model=model, concept=system_concept, name=system_data['nom'], defaults={'code': system_data['code'], 'description': system_data['description']})
                                slot, created = OSlot.objects.get_or_create(model=model, subject=system, predicate=system_etiquette_predicate, object=analyse_dsn_etiquette)

                                if system_data['etablissement']:
                                    etablissements = system_data['etablissement'].split('/')
                                    for etablissement_ in etablissements:
                                        etablissement, created = OInstance.objects.get_or_create(model=model, concept=etablissement_concept, name=etablissement_, defaults={'description': ''})
                                        slot, created = OSlot.objects.get_or_create(model=model, subject=system, predicate=system_etablissement_predicate, object=etablissement)

                                if system_data['manufacturier']:
                                    manufacturier, created = OInstance.objects.get_or_create(model=model, concept=manufacturier_concept, name=system_data['manufacturier'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=system, predicate=system_manufacturier_predicate, object=manufacturier)

                                for system_instance in similar_system_instances:
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=system_instance, predicate=system_instance_system_predicate, object=system)

                                
                        workbook.close()
                self.stdout.write(self.style.SUCCESS('Successfully updated model "%s"' % model_name))

            except Exception as e:
                #self.stdout.write(self.style.ERROR('Unable to update model "%s": %s' % (model_name, str(e))))
                raise CommandError('Unable to update model "%s": %s' % (model_name, str(e)))

    # def update_system(model, system_name):
    #     installation_name = system_name.name.split('_')[0]
    #     installation, created = OInstance.objects.get_or_create(model=model, concept=installation_concept, name=installation_name, defaults={'description': ''})
    #     slot, created = OSlot.objects.get_or_create(model=model, subject=system_instance, predicate=system_instance_installation_predicate, object=installation)
