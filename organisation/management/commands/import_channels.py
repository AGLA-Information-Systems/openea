import json
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from ontology.models import (OConcept, OInstance, OModel, OPredicate,
                             ORelation, OSlot)


class Command(BaseCommand):
    help = 'Create an model with its permissions'

    def add_arguments(self, parser):
        parser.add_argument('model_name', nargs=1, type=str)
        parser.add_argument('model_version', nargs=1, type=str)

    def handle(self, *args, **options):
        model_name = options['model_name'][0]
        model_version = options['model_version'][0]
        
        with transaction.atomic():
            try:
                model, created = OModel.objects.get_or_create(name=model_name, version=model_version)
                print(model.id)
                property_relation, created = ORelation.objects.get_or_create(model=model, name='a pour propriété', defaults={'description': ''})

                communication_source_format_relation, created = ORelation.objects.get_or_create(model=model, name='format à la source', defaults={'description': ''})
                communication_destination_format_relation, created = ORelation.objects.get_or_create(model=model, name='format à la destination', defaults={'description': ''})
                utilise_relation, created = ORelation.objects.get_or_create(model=model, name='utilise', defaults={'description': ''})
                transporte_relation, created = ORelation.objects.get_or_create(model=model, name='transporte', defaults={'description': ''})
                implemente_par_relation, created = ORelation.objects.get_or_create(model=model, name='est implémenté par', defaults={'description': ''})
                fait_partie_relation, created = ORelation.objects.get_or_create(model=model, name='fait partie de', defaults={'description': ''})

                flux_donnees_relation, created = ORelation.objects.get_or_create(model=model, name='flux de données', defaults={'description': ''})

                etablissement_concept, created = OConcept.objects.get_or_create(model=model, name='Établissement', defaults={'description': ''})
                installation_concept, created = OConcept.objects.get_or_create(model=model, name='Installation', defaults={'description': ''})
                communication_concept, created = OConcept.objects.get_or_create(model=model, name='Canal de communication', defaults={'description': ''})
                format_concept, created = OConcept.objects.get_or_create(model=model, name='Format de données', defaults={'description': ''})
                norme_concept, created = OConcept.objects.get_or_create(model=model, name='Norme/Standard', defaults={'description': ''})
                systeme_concept, created = OConcept.objects.get_or_create(model=model, name='Système informatique ou Outil', defaults={'description': ''})
                instance_systeme_concept, created = OConcept.objects.get_or_create(model=model, name='Instance de Système informatique', defaults={'description': ''})
                donnees_concept, created = OConcept.objects.get_or_create(model=model, name="Objet d'affaires", defaults={'description': ''})
                format_concept, created = OConcept.objects.get_or_create(model=model, name='Format', defaults={'description': ''})
                mode_concept, created = OConcept.objects.get_or_create(model=model, name='Mode de communication', defaults={'description': ''})

                etiquette_concept, created = OConcept.objects.get_or_create(model=model, name='Étiquette', defaults={'description': ''})

                installation_etablissement_predicate, created = OPredicate.objects.get_or_create(model=model, subject=installation_concept, relation=fait_partie_relation, object=etablissement_concept, defaults={'description': ''})

                communication_installation_predicate, created = OPredicate.objects.get_or_create(model=model, subject=communication_concept, relation=property_relation, object=installation_concept, defaults={'description': ''})
                communication_source_predicate, created = OPredicate.objects.get_or_create(model=model, subject=instance_systeme_concept, relation=flux_donnees_relation, object=communication_concept, defaults={'description': ''})
                communication_source_format_predicate, created = OPredicate.objects.get_or_create(model=model, subject=communication_concept, relation=communication_source_format_relation, object=format_concept, defaults={'description': ''})
                communication_destination_predicate, created = OPredicate.objects.get_or_create(model=model, subject=communication_concept, relation=flux_donnees_relation, object=instance_systeme_concept, defaults={'description': ''})
                communication_destination_format_predicate, created = OPredicate.objects.get_or_create(model=model, subject=communication_concept, relation=communication_destination_format_relation, object=format_concept, defaults={'description': ''})
                communication_systeme_predicate, created = OPredicate.objects.get_or_create(model=model, subject=communication_concept, relation=implemente_par_relation, object=instance_systeme_concept, defaults={'description': ''})
                format_norme_predicate, created = OPredicate.objects.get_or_create(model=model, subject=format_concept, relation=utilise_relation, object=norme_concept, defaults={'description': ''})
                communication_donnees_predicate, created = OPredicate.objects.get_or_create(model=model, subject=communication_concept, relation=transporte_relation, object=donnees_concept, defaults={'description': ''})
                communication_mode_predicate, created = OPredicate.objects.get_or_create(model=model, subject=communication_concept, relation=utilise_relation, object=mode_concept, defaults={'description': ''})

                instance_systeme_systeme_predicate, created = OPredicate.objects.get_or_create(model=model, subject=instance_systeme_concept, relation=property_relation, object=systeme_concept, defaults={'description': ''})
                instance_systeme_installation_predicate, created = OPredicate.objects.get_or_create(model=model, subject=instance_systeme_concept, relation=property_relation, object=installation_concept, defaults={'description': ''})
                instance_systeme_etiquette_predicate, created = OPredicate.objects.get_or_create(model=model, subject=instance_systeme_concept, relation=property_relation, object=etiquette_concept, defaults={'description': ''})
                systeme_etiquette_predicate, created = OPredicate.objects.get_or_create(model=model, subject=systeme_concept, relation=property_relation, object=etiquette_concept, defaults={'description': ''})
                
                document_dir = Path(r'/mnt/c/Data/cartographie')
                paths = [
                    'Interfaces.xlsx'
                ]
                for path in paths:
                    for xlsx_file in document_dir.glob(path):
                        # xlsx_file is a Path object
                        # if you use old libraries, you have to use str(xlsx_file) to convert the Path to a str
                        print(xlsx_file)

                        workbook = load_workbook(filename=xlsx_file, read_only=True)
                        #sheet = workbook.active
                        for sheet in workbook.worksheets:
                            print(sheet.title)
                            limit = sheet.max_row + 1
                            i = 2
                            data = []
                            while i < limit:
                                nom_canal = sheet['A' + str(i)].value
                                if nom_canal:
                                    print(nom_canal, ': ', i)
                                    communication_data = {
                                        'nom': sheet['A' + str(i)].value.strip(),
                                        'installation': sheet['C' + str(i)].value.strip() if sheet['C' + str(i)].value else None,
                                        'instance_source': sheet['D' + str(i)].value.strip() if sheet['D' + str(i)].value else None,
                                        'statut_instance_source': sheet['E' + str(i)].value.strip() if sheet['E' + str(i)].value else None,
                                        'instance_destination': sheet['F' + str(i)].value.strip() if sheet['F' + str(i)].value else None,
                                        'statut_instance_destination': sheet['G' + str(i)].value.strip() if sheet['G' + str(i)].value else None,
                                        'mode': sheet['J' + str(i)].value.strip() if sheet['J' + str(i)].value else None,
                                        'instance_interface': sheet['K' + str(i)].value.strip() if sheet['K' + str(i)].value else None,
                                        'source_format': sheet['L' + str(i)].value.strip() if sheet['L' + str(i)].value else None,
                                        'destination_format': sheet['M' + str(i)].value.strip() if sheet['M' + str(i)].value else None
                                    }
                                    data.append(communication_data)
                                i = i + 1

                            print(json.dumps(data, indent=4, ensure_ascii=False))

                            dsn_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='DSN', defaults={'description': ''})
                            cible_dsn_vitrine_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='Cible DSN Vitrine', defaults={'description': ''})
                            cible_dsn_etiquette, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name='Cible DSN', defaults={'description': ''})

                            for communication_data in data:
                                communication, created = OInstance.objects.get_or_create(model=model, concept=communication_concept, name=communication_data['nom'], defaults={'description': ''})

                                if communication_data['installation']:
                                    installation, created = OInstance.objects.get_or_create(model=model, concept=installation_concept, name=communication_data['installation'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=communication, predicate=communication_installation_predicate, object=installation)

                                if communication_data['instance_source']:
                                    instance_source, created = OInstance.objects.get_or_create(model=model, concept=instance_systeme_concept, name=communication_data['instance_source'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=instance_source, predicate=communication_source_predicate, object=communication)

                                    installation = Command.add_installation_from_instance_systeme(model=model, instance_systeme=instance_source, installation_concept=installation_concept, instance_systeme_installation_predicate=instance_systeme_installation_predicate)

                                    if communication_data['statut_instance_source']:
                                        statut_instance_source_ = communication_data['statut_instance_source']
                                    else:
                                        statut_instance_source_ = 'Inconnu'
                                    statut_instance_source, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name=statut_instance_source_, defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=instance_source, predicate=instance_systeme_etiquette_predicate, object=statut_instance_source)
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=instance_source, predicate=instance_systeme_etiquette_predicate, object=dsn_etiquette)

                                if communication_data['instance_destination']:
                                    instance_destination, created = OInstance.objects.get_or_create(model=model, concept=instance_systeme_concept, name=communication_data['instance_destination'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=communication, predicate=communication_destination_predicate, object=instance_destination)
                                    
                                    installation = Command.add_installation_from_instance_systeme(model=model, instance_systeme=instance_destination, installation_concept=installation_concept, instance_systeme_installation_predicate=instance_systeme_installation_predicate)

                                    if communication_data['statut_instance_destination']:
                                        statut_instance_destination_ = communication_data['statut_instance_destination']
                                    else:
                                        statut_instance_destination_ = 'Inconnu'
                                    statut_instance_destination, created = OInstance.objects.get_or_create(model=model, concept=etiquette_concept, name=statut_instance_destination_, defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=instance_destination, predicate=instance_systeme_etiquette_predicate, object=statut_instance_destination)
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=instance_destination, predicate=instance_systeme_etiquette_predicate, object=dsn_etiquette)
                                
                                if communication_data['source_format']:
                                    source_format, created = OInstance.objects.get_or_create(model=model, concept=format_concept, name=communication_data['source_format'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=communication, predicate=communication_source_format_predicate, object=source_format)

                                    norme, created = OInstance.objects.get_or_create(model=model, concept=norme_concept, name=communication_data['source_format'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=source_format, predicate=format_norme_predicate, object=norme)

                                if communication_data['destination_format']:
                                    destination_format, created = OInstance.objects.get_or_create(model=model, concept=format_concept, name=communication_data['destination_format'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=communication, predicate=communication_destination_format_predicate, object=destination_format)

                                    norme, created = OInstance.objects.get_or_create(model=model, concept=norme_concept, name=communication_data['destination_format'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=destination_format, predicate=format_norme_predicate, object=norme)

                                if communication_data['mode']:
                                    mode, created = OInstance.objects.get_or_create(model=model, concept=mode_concept, name=communication_data['mode'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=communication, predicate=communication_mode_predicate, object=mode)

                                if communication_data['instance_interface']:
                                    instance_interface, created = OInstance.objects.get_or_create(model=model, concept=instance_systeme_concept, name=communication_data['instance_interface'], defaults={'description': ''})
                                    slot, created = OSlot.objects.get_or_create(model=model, subject=communication, predicate=communication_systeme_predicate, object=instance_interface)


                        workbook.close()
                self.stdout.write(self.style.SUCCESS('Successfully updated model "%s"' % model_name))

            except Exception as e:
                #self.stdout.write(self.style.ERROR('Unable to update model "%s": %s' % (model_name, str(e))))
                raise CommandError('Unable to update model "%s": %s' % (model_name, str(e)))

    def add_installation_from_instance_systeme(model, instance_systeme, installation_concept, instance_systeme_installation_predicate):
        installation_name = instance_systeme.name.split('_')[0]
        installation, created = OInstance.objects.get_or_create(model=model, concept=installation_concept, name=installation_name, defaults={'description': ''})
        slot, created = OSlot.objects.get_or_create(model=model, subject=instance_systeme, predicate=instance_systeme_installation_predicate, object=installation)